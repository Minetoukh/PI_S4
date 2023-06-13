
from django.contrib import messages
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password, check_password 
from django.urls import reverse_lazy
from .models import *
from .forms import *
from django.shortcuts import render
from django.views.generic import UpdateView, DeleteView, CreateView
import pandas as pd

from django.shortcuts import render
# Create your views here.
from django.conf import settings
def index(request):
    if request.session.get('username', None):
        x = request.session['username']
        return render(request, 'index.html', {"name": x})
    else:
        return redirect('login')


 


def login(request):
    if request.session.get('username', None):
        return redirect('index')
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        if not len(username):
            messages.warning(request, "username is empty")
            redirect('login')
        elif not len(password):
            messages.warning(request, "password is empty")
            redirect('login')
        else:
            pass
        if User.objects.filter(username=username):
            user = User.objects.filter(username=username)[0]
            password_hash = user.password
            resp = check_password(password, password_hash)
            if resp == 1:
                request.session['id'] = user.id
                request.session['username'] = username
                return redirect('index')
            else:
                messages.warning(request, "username or password is incorrect")
                redirect('login')
        else:
            messages.warning(request, "account 404")
            redirect('login')
    return render(request, 'login.html', {})


# Logs
def logout(request):
    for key in list(request.session.keys()):
        del request.session[key]
    return redirect('login')

def register(request):
    if request.session.get('username', None):
        return redirect('index')
    if request.method == "POST":
        username = request.POST['username']
        email = request.POST['email']
        if User.objects.filter(username=username) or User.objects.filter(email=email):
            messages.warning(request, "account already exist")
        else:
            password = request.POST['password']
            error = []
            if (len(username) < 3):
                error.append(1)
                messages.warning(
                    request, "Username Field must be greater than 3 character.")
            if (len(password) < 5):
                error.append(1)
                messages.warning(
                    request, "Password Field must be greater than 5 character.")
            if (len(email) == 0):
                error.append(1)
                messages.warning(request, "Email field can't be empty")
            if (len(error) == 0):
                password_hash = make_password(password)
                user = User(username=username,
                            password=password_hash, email=email)
                user.save()
                messages.info(request, "account created successfully")
                redirect('register')
            else:
                redirect('register')
    return render(request, 'register.html', {})

#dep
def dep(request):
    
    dep = Dep.objects.all()
    return render(request, 'dep.html', {"dep": dep})


class depCreateView(CreateView):
    model = Dep
    template_name = 'dep_form.html'
    fields = '__all__'
    success_url = reverse_lazy('dep')


class depUpdateView(UpdateView):
    model = Dep
    form_class = depForm
    template_name = 'update_dep.html'
    success_url = reverse_lazy('dep')


class depDeleteView(DeleteView):
    model = Dep
    template_name = 'dep_confirm_delete.html'
    success_url = reverse_lazy('dep')



#entreprise
def entre(request):
    
    entre = Entre.objects.all()
    return render(request, 'entre.html', {"entre": entre})


class entreCreateView(CreateView):
    model = Entre
    template_name = 'entre_form.html'
    fields = '__all__'
    success_url = reverse_lazy('entre')


class entreUpdateView(UpdateView):
    model = Entre
    form_class = entreForm
    template_name = 'update_entre.html'
    success_url = reverse_lazy('entre')


class entreDeleteView(DeleteView):
    model = Entre
    template_name = 'entre_confirm_delete.html'
    success_url = reverse_lazy('entre')



#et
def et(request):
    dep=Dep.objects.all()
    etud =Etud.objects.all()
    return render(request,'etud_form.html' ,{"etud": etud,"dep": dep})
#etud
def etud(request):
     
    etud = Etud.objects.all()
    return render(request, 'etud.html', {"etud": etud})



class etudCreateView(CreateView):
    model = Etud
    model = Dep
    template_name = 'etud_form.html'
    fields = '__all__'
    success_url = reverse_lazy('etud')

from django.urls import reverse_lazy
from django.views.generic.edit import UpdateView
from .models import Etud
from .forms import etudForm
class etudUpdateView(UpdateView):
    model = Etud
    form_class = etudForm
    template_name = 'update_etud.html'

    def get_success_url(self):
        dep = self.object.Dep
        success_url = reverse_lazy('etud') + f'?dep={dep}'
        return success_url


class etudDeleteView(DeleteView):
    model = Etud
    template_name = 'etud_confirm_delete.html'
    success_url = reverse_lazy('etud')    

#encours
def encours(request):
    return render(request, 'encours.html')    





#EN_type
def type_Encadrant(request):
    
    type_Encadrant = Type_Encadrant.objects.all()
    return render(request, 'Type_Encadrant.html', {"type_Encadrant": type_Encadrant})


class Type_EncadrantCreateView(CreateView):
    model = Type_Encadrant
    template_name = 'Type_Encadrant_form.html'
    fields = '__all__'
    success_url = reverse_lazy('type_Encadrant')


class Type_EncadrantUpdateView(UpdateView):
    model = Type_Encadrant
    form_class = Type_EncadrantForm
    template_name = 'update_Type_Encadrant.html'
    success_url = reverse_lazy('type_Encadrant')


class Type_EncadrantDeleteView(DeleteView):
    model = Type_Encadrant
    template_name = 'Type_Encadrant_confirm_delete.html'
    success_url = reverse_lazy('type_Encadrant') 




#Type_stage
def type_stage(request):
    
    type_stage = Type_stage.objects.all()
    return render(request, 'type_stage.html', {"type_stage": type_stage})


class type_stageCreateView(CreateView):
    model = Type_stage
    template_name = 'type_stage_form.html'
    fields = '__all__'
    success_url = reverse_lazy('type_stage')


class type_stageUpdateView(UpdateView):
    model = Type_stage
    form_class = Type_stageForm
    template_name = 'update_type_stage.html'
    success_url = reverse_lazy('type_stage')


class type_stageDeleteView(DeleteView):
    model = Type_stage
    template_name = 'type_stage_confirm_delete.html'
    success_url = reverse_lazy('type_stage') 






#Simester
def simester(request):
    
    simester = Simester.objects.all()
    return render(request, 'simester.html', {"simester": simester})


class simesterCreateView(CreateView):
    model = Simester
    template_name = 'simester_form.html'
    fields = '__all__'
    success_url = reverse_lazy('simester')


class simesterUpdateView(UpdateView):
    model = Simester
    form_class = simesterForm
    template_name = 'update_simester.html'
    success_url = reverse_lazy('simester')


class simesterDeleteView(DeleteView):
    model = Simester
    template_name = 'simester_confirm_delete.html'
    success_url = reverse_lazy('simester') 





#encadrant
def encadrant(request):
    
    encadrant = Encadrent.objects.all()
    return render(request, 'Encadrent.html', {"encadrant": encadrant})


class encadrantCreateView(CreateView):
    model = Encadrent
    template_name = 'Encadrent_form.html'
    fields = '__all__'
    success_url = reverse_lazy('encadrant')


class encadrantUpdateView(UpdateView):
    model = Encadrent
    form_class = encadrantForm
    template_name = 'update_Encadrent.html'
    success_url = reverse_lazy('encadrant')


class encadrantDeleteView(DeleteView):
    model = Encadrent
    template_name = 'Encadrent_confirm_delete.html'
    success_url = reverse_lazy('encadrant')    



#etd
def etd(request):
    return render(request, 'etd.html')

#encad
def encad(request):
    return render(request, 'ecd.html')

#stg
def stg(request):
    return render(request, 'stg.html')

 
 

def ajouter_groupe(request):
    if request.method == 'POST':
        Nom = request.POST['Nom']
        prenom = request.POST['prenom']
        email = request.POST['email']
        filiere = request.POST['filiere']
        libele = request.POST['libele']
        groupe = Groupe(Nom=Nom, prenom=prenom, email=email, filiere=filiere, libele=libele)
        groupe.save()
        return render(request, 'encours.html')
    else:
        return render(request, 'etd.html')
    

def ajouter_stage(request):
    if request.method == 'POST':
        titre = request.POST['titre']
        dattedebut = request.POST['dattedebut']
        dattefin = request.POST['dattefin']
        datesoutenance = request.POST['datesoutenance']
        lieu = request.POST['lieu']
        Annee = request.POST['Annee']
        stage = Stage(titre=titre, dattedebut=dattedebut, dattefin=dattefin, datesoutenance=datesoutenance, lieu=lieu, Annee=Annee)
        stage.save()
        return render(request, 'encours.html')
    else:
        return render(request, 'stg.html')



def ajouter_encd(request):
    if request.method == 'POST':
        nom = request.POST['nom']
        prenom = request.POST['prenom']
        email = request.POST['email']
        profile = request.POST['profile']
        
        encd = Encadrent(nom=nom, prenom=prenom, email=email, profile=profile)
        encd.save()
        return render(request, 'encours.html')
    else:
        return render(request, 'ecd.html')
    
#stage
def stage(request):
    
    stage = Stage.objects.all()
    return render(request, 'stage_encours.html', {"stage": stage})


from django.shortcuts import render, redirect
from django.http import JsonResponse
from .forms import grpForm
from .models import Etud

def creer_groupe(request):
    if request.method == 'POST':
        idSimester = request.POST.get('idSimester')
        idEtud = request.POST.get('idEtud')
        libele = request.POST.get('libele')
        membres = request.POST.getlist('membres')
        groupe = grp(libele=libele, idSimester=idSimester, idEtud=idEtud)
        groupe.save()
        groupe.membres.set(membres)
        return JsonResponse({'encours': True})
    else:
        return render(request, 'etd.html')


def autocomplete_etudiants(request):
    term = request.POST.get('matricule')
    etudiants = Etud.objects.filter(matricule__startswith=term).values('matricule', 'Nom', 'prenom')
    data = [{'matricule': etudiant['matricule'], 'NomComplet': str(etudiant['matricule']) +"- "+etudiant['Nom'] + ' ' + etudiant['prenom']} for etudiant in etudiants]
    return JsonResponse(data, safe=False)





import pandas as pd
from django.shortcuts import render
 

def import_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['file']
        data_frame = pd.read_excel(excel_file)

        # Parcours des lignes du DataFrame pandas
        for index, row in data_frame.iterrows():
            nom = row['nom']
            prenom = row['prenom']
            email = row['email']
            numero = row['numero']

            # Création d'une instance Person et enregistrement dans la base de données
            Encadrent.objects.create(nom=nom, prenom=prenom, email=email, numero=numero)

        return render(request, 'import_success.html')

    return render(request, 'import_excel.html')


 

def view_encadrent(request):
    encadrent = Encadrent.objects.all()
    return render(request, 'view_encadrent.html', {'encadrent': encadrent})

 

import openpyxl
from django.http import HttpResponse

def export_excel(request):
    # Créer un nouveau classeur Excel
    workbook = openpyxl.Workbook()
    
    # Sélectionner la feuille active
    sheet = workbook.active
    
    # Ajouter des en-têtes de colonne
    sheet['A1'] = 'Nom'
    sheet['B1'] = 'Prénom'
    sheet['C1'] = 'Email'
    sheet['D1'] = 'Numéro'
    
    # Obtenir les données à exporter (supposons que vous ayez un modèle "Person" avec les champs nom, prenom, email, et numero)
    queryset = Encadrent.objects.all()
    
    # Parcourir les données et les ajouter à la feuille Excel
    row = 2  # Commencer à la deuxième ligne
    for Encadrent in queryset:
        sheet.cell(row=row, column=1).value = Encadrent.nom
        sheet.cell(row=row, column=2).value = Encadrent.prenom
        sheet.cell(row=row, column=3).value = Encadrent.email
        sheet.cell(row=row, column=4).value = Encadrent.numero
        row += 1
    
    # Définir le type de contenu de la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Définir le nom du fichier Excel
    response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
    
    # Sauvegarder le classeur Excel dans la réponse HTTP
    workbook.save(response)
    
    return response

from django.shortcuts import render
from .models import Etud, Dep

def add_etudiant(request):
    if request.method == 'POST':
        matricule = request.POST.get('matricule')
        Nom = request.POST.get('Nom')
        prenom = request.POST.get('prenom')
        dep = request.POST.get('dep')

        # Récupérer le département correspondant à l'ID sélectionné
        dep = Dep.objects.get(id=dep)

        # Créer une nouvelle instance d'Etud avec les données fournies
        etud = Etud(matricule=matricule, Nom=Nom, prenom=prenom, Dep=dep)
        etud.save()
        

        # Rediriger vers une page de confirmation ou effectuer toute autre action souhaitée

    # Récupérer tous les départements disponibles
    dep = Dep.objects.all()

    context = {
        'dep': dep
    }
    etud = Etud.objects.all()
    return render(request, 'etud.html', {"etud": etud})



